#!/usr/bin/env python3
"""
Load curation status into the REFERENCE table.

This script reads curation status information from a file and updates
the CURATION_STATUS column in the REFERENCE table.

Input formats supported:
- Tab-delimited text file
- Excel file (.xlsx, .xls)

File format (columns):
- Column 0: Gene name (optional)
- Column 2: PubMed ID
- Column 3: Curation status

Status values are prioritized as:
1. "abs get text"
2. "abs"
3. "full"
4. "unlink" -> becomes "not gene specific" or "High Priority Uncurated"

Original Perl: loadCurationStatus.pl
Converted to Python: 2024
"""

import argparse
import logging
import os
import sys
from collections import defaultdict
from pathlib import Path

from dotenv import load_dotenv
from sqlalchemy import and_
from sqlalchemy.orm import Session

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from cgd.db.engine import SessionLocal
from cgd.models.models import Reference

load_dotenv()

logger = logging.getLogger(__name__)

# Status priority order (first has highest priority)
STATUS_PRIORITY = ["abs get text", "abs", "full", "unlink"]


def setup_logging(log_file: Path = None, verbose: bool = False) -> None:
    """Configure logging."""
    level = logging.DEBUG if verbose else logging.INFO
    handlers = [logging.StreamHandler()]

    if log_file:
        log_file.parent.mkdir(parents=True, exist_ok=True)
        handlers.append(logging.FileHandler(log_file))

    logging.basicConfig(
        level=level,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=handlers,
    )


def read_excel_file(filepath: Path, sheet_name: str = None) -> list[list]:
    """
    Read data from an Excel file.

    Args:
        filepath: Path to Excel file
        sheet_name: Sheet name to read (default: first sheet or named sheet)

    Returns:
        List of rows (each row is a list of cell values)
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        raise

    wb = openpyxl.load_workbook(filepath, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=2):  # Skip header
        row_data = [cell.value for cell in row]
        rows.append(row_data)

    return rows


def read_text_file(filepath: Path, delimiter: str = "\t") -> list[list]:
    """
    Read data from a text file.

    Args:
        filepath: Path to text file
        delimiter: Column delimiter

    Returns:
        List of rows (each row is a list of cell values)
    """
    rows = []

    with open(filepath) as f:
        for line_num, line in enumerate(f, 1):
            if line_num == 1:  # Skip header
                continue

            line = line.strip()
            if not line:
                continue

            parts = line.split(delimiter)
            rows.append(parts)

    return rows


def parse_curation_data(rows: list[list]) -> tuple[dict, set]:
    """
    Parse curation data from rows.

    Args:
        rows: List of data rows

    Returns:
        Tuple of (data dict mapping pubmed -> set of statuses,
                  set of pubmeds with uncurated genes)
    """
    data = defaultdict(set)
    uncurated_genes = set()

    for row in rows:
        # Ensure we have enough columns
        if len(row) < 4:
            continue

        gene = row[0]
        pubmed = row[2]
        status = row[3] if len(row) > 3 else None

        # Skip if no pubmed
        if not pubmed:
            continue

        # Try to convert pubmed to int
        try:
            pubmed = int(pubmed)
        except (ValueError, TypeError):
            continue

        # Handle missing status
        if not status or str(status).strip() == "":
            # If we have a gene but no status, track it
            if gene and str(gene).strip() != "":
                uncurated_genes.add(pubmed)
            continue

        status = str(status).strip()

        # Skip uninformative statuses
        if status == "ref_bad":
            continue

        data[pubmed].add(status)

    return dict(data), uncurated_genes


def determine_final_status(
    statuses: set[str],
    pubmed: int,
    uncurated_genes: set[int],
) -> str | None:
    """
    Determine the final curation status based on priority.

    Args:
        statuses: Set of status values for this pubmed
        pubmed: PubMed ID
        uncurated_genes: Set of pubmeds with uncurated genes

    Returns:
        Final status string or None
    """
    # If multiple statuses, use priority order
    if len(statuses) > 1:
        for priority_status in STATUS_PRIORITY:
            if priority_status in statuses:
                return priority_status
    elif "unlink" in statuses:
        # Special handling for unlink
        if pubmed in uncurated_genes:
            return "High Priority Uncurated"
        else:
            return "not gene specific"

    # Return the single status if only one
    if len(statuses) == 1:
        return next(iter(statuses))

    return None


def load_curation_status(
    session: Session,
    data: dict,
    uncurated_genes: set[int],
    created_by: str,
) -> dict:
    """
    Load curation status into the database.

    Args:
        session: Database session
        data: Dictionary mapping pubmed -> set of statuses
        uncurated_genes: Set of pubmeds with uncurated genes
        created_by: User performing the update

    Returns:
        Dictionary with statistics
    """
    stats = {
        "pubmeds_processed": 0,
        "references_updated": 0,
        "references_not_found": 0,
    }

    for pubmed, statuses in sorted(data.items()):
        stats["pubmeds_processed"] += 1

        # Determine final status
        final_status = determine_final_status(statuses, pubmed, uncurated_genes)

        if not final_status:
            continue

        logger.debug(f"PubMed {pubmed}: {final_status} (from {statuses})")

        # Find reference by pubmed
        reference = session.query(Reference).filter(
            Reference.pubmed == pubmed
        ).first()

        if not reference:
            logger.debug(f"Reference not found for PubMed {pubmed}")
            stats["references_not_found"] += 1
            continue

        # Update curation status
        reference.curation_status = final_status
        stats["references_updated"] += 1

    return stats


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Load curation status into the REFERENCE table"
    )
    parser.add_argument(
        "input_file",
        type=Path,
        help="Input file (Excel or tab-delimited text)",
    )
    parser.add_argument(
        "--created-by",
        default=os.getenv("DB_USER", "SCRIPT"),
        help="Database user name for tracking",
    )
    parser.add_argument(
        "--sheet-name",
        default="Reference bookkeeping, comments",
        help="Excel sheet name (for .xls/.xlsx files)",
    )
    parser.add_argument(
        "--log-file",
        type=Path,
        help="Path to log file",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Enable verbose output",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse file but don't modify database",
    )

    args = parser.parse_args()

    # Setup logging
    setup_logging(args.log_file, args.verbose)

    # Validate input file
    if not args.input_file.exists():
        logger.error(f"Input file not found: {args.input_file}")
        sys.exit(1)

    logger.info(f"Input file: {args.input_file}")
    logger.info(f"Created by: {args.created_by}")

    # Read input file
    try:
        suffix = args.input_file.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            logger.info(f"Reading Excel file, sheet: {args.sheet_name}")
            rows = read_excel_file(args.input_file, args.sheet_name)
        else:
            logger.info("Reading tab-delimited text file")
            rows = read_text_file(args.input_file)
    except Exception as e:
        logger.error(f"Error reading input file: {e}")
        sys.exit(1)

    logger.info(f"Read {len(rows)} data rows")

    # Parse data
    data, uncurated_genes = parse_curation_data(rows)
    logger.info(f"Found {len(data)} unique PubMed IDs")
    logger.info(f"Found {len(uncurated_genes)} PubMeds with uncurated genes")

    if args.dry_run:
        logger.info("DRY RUN - no database modifications")
        for pubmed, statuses in sorted(data.items()):
            final_status = determine_final_status(statuses, pubmed, uncurated_genes)
            if final_status:
                logger.info(f"Would update PubMed {pubmed}: {final_status}")
        return

    try:
        with SessionLocal() as session:
            stats = load_curation_status(
                session,
                data,
                uncurated_genes,
                args.created_by,
            )

            session.commit()
            logger.info("Transaction committed successfully")

            logger.info("=" * 50)
            logger.info("Load Summary:")
            logger.info(f"  PubMeds processed: {stats['pubmeds_processed']}")
            logger.info(f"  References updated: {stats['references_updated']}")
            logger.info(f"  References not found: {stats['references_not_found']}")
            logger.info("=" * 50)

    except Exception as e:
        logger.error(f"Error loading curation status: {e}")
        logger.exception("Full traceback:")
        sys.exit(1)


if __name__ == "__main__":
    main()
